from django.contrib import admin, messages
from django.db import models
from django.forms import widgets
from django.http import HttpResponse, StreamingHttpResponse, FileResponse
from django.utils import timezone
from django.conf import settings
from django.core.signals import request_finished, request_started
from django.dispatch import receiver
from .models import Questionnaire, Resource, Question, Choice, AnswerSheet, TextAnswer, FileAnswer

import os
import zipfile
import openpyxl
from threading import Timer
import tempfile


class QuestionInline(admin.TabularInline):
    model = Question
    show_change_link = True
    extra = 0

    def has_change_permission(self, request, obj=None):
        return not obj or not obj.status and super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        return not obj or not obj.status and super().has_delete_permission(request, obj)


class ResourceInline(admin.TabularInline):
    model = Resource
    extra = 0

    def has_change_permission(self, request, obj=None):
        return not obj or not obj.status and super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        return not obj or not obj.status and super().has_delete_permission(request, obj)


@admin.register(Questionnaire)
class QuestionnaireAdmin(admin.ModelAdmin):
    list_display = ['name', 'status', 'timePublish', 'deadLine',
                    'numQuestion', 'numAnswerSheet']
    list_filter = ['status', 'timePublish', 'deadLine']

    inlines = [QuestionInline, ResourceInline]

    def has_change_permission(self, request, obj=None):
        return not obj or not obj.status and super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        return not obj or not obj.status and super().has_delete_permission(request, obj)

    def save_model(self, request, obj, form, change):
        t = timezone.now()
        obj.userCreate = obj.userCreate or request.user
        obj.userModify = request.user
        obj.timeCreate = obj.timeCreate or t
        obj.timeModify = t
        if obj.status:
            if not obj.timePublish:
                self.message_user(request, '未指定发布时间，不能发布。', messages.ERROR)
                obj.status = False
            else:
                Questionnaire.objects.update(status=False)
                obj.status = True
        super().save_model(request, obj, form, change)

    def save_formset(self, request, form, formset, change):
        if formset.model is Question:
            objs = formset.save(commit=False)
            for obj in formset.deleted_objects:
                obj.delete()
            for obj in objs:
                t = timezone.now()
                obj.userCreate = obj.userCreate or request.user
                obj.userModify = request.user
                obj.timeCreate = obj.timeCreate or t
                obj.timeModify = t
                obj.save()
        elif formset.model is Resource:
            objs = formset.save(commit=False)
            for obj in formset.deleted_objects:
                obj.delete()
            for obj in objs:
                t = timezone.now()
                obj.userCreate = obj.userCreate or request.user
                obj.userModify = request.user
                obj.timeCreate = obj.timeCreate or t
                obj.timeModify = t
                obj.save()
        else:
            super().save_formset(request, form, formset, change)

    def delete_queryset(self, request, queryset):
        for obj in queryset:
            if self.has_delete_permission(request, obj):
                obj.delete()


class ChoiceInline(admin.TabularInline):
    model = Choice
    extra = 0


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['questionnaire', 'order', 'type', 'require', 'description']
    list_filter = ['questionnaire']

    inlines = [ChoiceInline]
    fields = ['questionnaire', 'order', 'type', 'description']
    readonly_fields = ['questionnaire']

    def has_module_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return not obj or not obj.questionnaire.status and super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        return not obj or not obj.questionnaire.status and super().has_delete_permission(request, obj)

    def save_model(self, request, obj, form, change):
        t = timezone.now()
        obj.userCreate = obj.userCreate or request.user
        obj.userModify = request.user
        obj.timeCreate = obj.timeCreate or t
        obj.timeModify = t
        super().save_model(request, obj, form, change)


class TextAnswerInline(admin.TabularInline):
    model = TextAnswer

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request, obj):
        return False


class FileAnswerInline(admin.TabularInline):
    model = FileAnswer

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request, obj):
        return False


def clear_file(fileName):
    try:
        os.remove(fileName)
        print('Successfully delete %s.' % fileName)
    except IOError:
        Timer(1, clear_file, [fileName]).start()
        print('Wait for another time.')


@admin.register(AnswerSheet)
class AnswerSheetAdmin(admin.ModelAdmin):
    list_display = ['email', 'timeSubmit', 'questionnaire']
    list_select_related = ['questionnaire']
    list_filter = ['timeSubmit', 'questionnaire']

    actions = ['reset_queryset', 'download_resources_queryset', 'download_excel_queryset']

    fields = ['email', 'timeSubmit', 'questionnaire']
    readonly_fields = ['email', 'timeSubmit', 'questionnaire']
    inlines = [TextAnswerInline, FileAnswerInline]

    def delete_queryset(self, request, queryset):
        for obj in queryset:
            if self.has_delete_permission(request, obj):
                obj.delete()

    def reset_queryset(self, request, queryset):
        queryset.update(password=None)
    reset_queryset.short_description = '重置密码'
    reset_queryset.allowed_permissions = ('change', )

    def download_resources_queryset(self, request, queryset):
        if queryset.exists():
            z_filePath = os.path.join(settings.MEDIA_ROOT,
                                      'enroll',
                                      timezone.localtime().strftime('%Y%m%d%H%M%S%f') + '.zip')
            z_file = zipfile.ZipFile(z_filePath, 'w')
            for obj in queryset:    # type: AnswerSheet
                for fileanswer in obj.fileanswer_set.all():   # type: FileAnswer
                    filePath = os.path.join(settings.MEDIA_ROOT, fileanswer.answer.name)
                    arcPath = os.path.join('questionnaire%d' % obj.questionnaire_id,
                                           '%s' % obj.email,
                                           'question%d' % fileanswer.question.order,
                                           os.path.split(fileanswer.answer.name)[-1])
                    z_file.write(filePath, arcPath)
            z_file.close()
            response = FileResponse(open(z_filePath, 'rb'), as_attachment=True, filename='export.zip')
            clear_file(z_filePath)
            return response
    download_resources_queryset.short_description = '下载附件'
    download_resources_queryset.allowed_permissions = ('view', )

    def download_excel_queryset(self, request, queryset):
        if queryset.exists():
            workBook = openpyxl.Workbook()
            del workBook['Sheet']
            for answerSheet in queryset:    # type: AnswerSheet
                sheetName = answerSheet.questionnaire.name
                if sheetName in workBook.sheetnames:
                    workSheet = workBook[sheetName]
                else:
                    workSheet = workBook.create_sheet(sheetName)
                    head = ['邮箱'] + list(answerSheet.textanswer_set.values_list('question__description', flat=True))
                    workSheet.append(head)
                data = [answerSheet.email] + list(answerSheet.textanswer_set.values_list('answer', flat=True))
                workSheet.append(data)
            filePath = os.path.join(settings.MEDIA_ROOT,
                                    'enroll',
                                    timezone.localtime().strftime('%Y%m%d%H%M%S%f') + '.xlsx')
            workBook.save(filePath)
            response = FileResponse(open(filePath, 'rb'), as_attachment=True, filename='export.xlsx')
            clear_file(filePath)
            return response
    download_excel_queryset.short_description = '下载表格'
    download_excel_queryset.allowed_permissions = ('view', )
